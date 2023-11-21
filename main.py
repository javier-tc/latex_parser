import re
import openpyxl
from openpyxl import load_workbook
from lxml import etree
from copy import copy

def parse_latex(latex_file):
    print("Parsing LaTeX file...")
    # Lee el archivo LaTeX
    with open(latex_file, 'r', encoding='utf-8') as file:
        latex_content = file.read()
    
    #claves
    clave_regex = r'begin{key} \w'
    claves = []
    for n in re.findall(clave_regex, latex_content):
        claves.append(n.split('begin{key} ')[-1])
    
    #eje
    eje_regex = r'Eje Temático:.*'
    ejes = []
    for n in re.findall(eje_regex, latex_content):
        ejes.append(n.split('Eje Temático: ')[-1].replace('\\\\',''))
        
    #contenido
    contenido_regex= r'Contenido:.*'
    contenidos = []
    for n in re.findall(contenido_regex, latex_content):
        contenidos.append(n.split('Contenido: ')[-1].replace('\\\\',''))
        
    #habilidad
    habilidad_regex = r'Habilidad:.*'
    habilidades = []
    for n in re.findall(habilidad_regex,latex_content):
        habilidades.append(n.split('Habilidad: ')[-1].replace('\\\\',''))
    
    #dificultad
    dificultad_regex = r'Dificultad:.*'
    dificultades = []
    for n in re.findall(dificultad_regex,latex_content):
        dificultades.append(n.split('Dificultad: ')[-1].replace('\\\\',''))
    
    n_preg = [n for n in range(1,66)]
    descripciones = ['' for n in range(1,66)]
    cobertura = ['' for n in range(1,66)]
    form_preg = ['' for n in range(1,66)]
    
    #print('claves:', len(claves), ', ejes: ', len(ejes), ', conten: ', len(contenidos), ', habil:', len(habilidades), ', dif:', len(dificultades))

    # Retorna los datos extraídos
    return [n_preg,claves,ejes,contenidos,descripciones,habilidades,cobertura, form_preg, dificultades]

def write_to_excel(data, excel_file):
    print("Writing to Excel...")
    # Crea un archivo de Excel
    temp_wb = load_workbook(excel_file, read_only=True)
    temp_ws = temp_wb.active # insert at the end (default)

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    
    for col_num, cell in enumerate(temp_ws[1], start=1):
        new_cell = new_sheet.cell(row=1, column=col_num, value=cell.value)
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.alignment = copy(cell.alignment)
        
    # Escribe los datos en la hoja de Excel
    for col_num, col_data in enumerate(data, start=1):
        # Itera sobre los elementos de la lista y escribe cada elemento en una fila
        for row_num, cell_value in enumerate(col_data, start=2):
            new_cell = new_sheet.cell(row=row_num, column=col_num, value=cell_value)

            new_cell.font = copy(temp_ws.cell(row=row_num, column=col_num).font)
            new_cell.border = copy(temp_ws.cell(row=row_num, column=col_num).border)
            new_cell.fill = copy(temp_ws.cell(row=row_num, column=col_num).fill)
            new_cell.number_format = copy(temp_ws.cell(row=row_num, column=col_num).number_format)
            new_cell.alignment = copy(temp_ws.cell(row=row_num, column=col_num).alignment)
    
    # Ajusta automáticamente el ancho de las columnas al tamaño del texto
    for column in new_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        new_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    # Guarda el nuevo archivo de Excel
    new_workbook.save('nuevo_excel.xlsx')

if __name__ == "__main__":
    # Especifica el archivo LaTeX y el archivo de Excel
    latex_file = "main.tex"
    excel_file = "HISEO-08.xlsx"

    # Parsea el archivo LaTeX
    data = parse_latex(latex_file)

    # Escribe los datos en el archivo Excel
    write_to_excel(data, excel_file)
    
    print("Proceso completo.")
    print("Press any key to exit")
    input()
    print("You pressed a key. Exiting now...")